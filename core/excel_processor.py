"""
Excel智能处理模块
支持：
1. 自动跳过无效行（注释、标题等）
2. 单表头/多表头自动识别
3. 可选调用LLM进行智能分析
4. 合并单元格处理
5. 列结构元数据生成
"""

import pandas as pd
import json
import re
import os
import requests
import logging
from openpyxl import load_workbook
from typing import Tuple, List, Dict, Optional, Any
from collections import defaultdict
from dataclasses import dataclass, asdict, field
from pathlib import Path

# 配置日志
logger = logging.getLogger(__name__)

# 导入配置（避免循环导入，使用延迟导入）

from .config import EXCEL_LLM_API_KEY, EXCEL_LLM_BASE_URL, EXCEL_LLM_MODEL, EXCEL_MAX_ROWS_PREVIEW, EXCEL_MAX_COLS_PREVIEW



@dataclass
class HeaderAnalysis:
    """表头分析结果"""
    skip_rows: int          # 需要跳过的无效行数
    header_rows: int        # 表头占用的行数
    header_type: str        # 'single' 或 'multi'
    data_start_row: int     # 数据开始行（1-indexed）
    confidence: str         # 置信度: high/medium/low
    reason: str             # 分析原因说明
    valid_cols: Optional[List[int]] = None  # 有效列的索引列表（1-indexed），None表示所有列都有效
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        result = asdict(self)
        if result.get('valid_cols') is None:
            result['valid_cols'] = None
        return result


@dataclass
class ExcelProcessResult:
    """Excel处理结果"""
    success: bool
    header_analysis: Optional[HeaderAnalysis]
    processed_file_path: Optional[str]      # 处理后的CSV文件路径
    metadata_file_path: Optional[str]       # 元数据JSON文件路径
    column_names: List[str]                 # 列名列表
    column_metadata: Dict[str, Dict]        # 列结构元数据
    row_count: int                          # 数据行数
    error_message: Optional[str]            # 错误信息
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        return {
            "success": self.success,
            "header_analysis": self.header_analysis.to_dict() if self.header_analysis else None,
            "processed_file_path": self.processed_file_path,
            "metadata_file_path": self.metadata_file_path,
            "column_names": self.column_names,
            "column_metadata": self.column_metadata,
            "row_count": self.row_count,
            "error_message": self.error_message
        }


class SmartHeaderProcessor:
    """智能表头处理器"""
    
    def __init__(self, filepath: str, sheet_name: str = None):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.wb = load_workbook(filepath, data_only=True)
        self.ws = self.wb[sheet_name] if sheet_name else self.wb.active
        self.merged_cells_map = self._build_merged_cells_map()
    
    def _build_merged_cells_map(self) -> Dict[Tuple[int, int], str]:
        """构建合并单元格映射"""
        merged_map = {}
        for merged_range in self.ws.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            value = self.ws.cell(min_row, min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_map[(row, col)] = value
        return merged_map
    
    def get_cell_value(self, row: int, col: int) -> Any:
        """获取单元格值，处理合并单元格"""
        if (row, col) in self.merged_cells_map:
            return self.merged_cells_map[(row, col)]
        return self.ws.cell(row, col).value
    
    def get_preview_data(self, max_rows: int = 15, max_cols: int = 10) -> List[List[Any]]:
        """获取预览数据用于分析"""
        actual_max_col = min(self.ws.max_column, max_cols)
        actual_max_row = min(self.ws.max_row, max_rows)
        
        data = []
        for row in range(1, actual_max_row + 1):
            row_data = []
            for col in range(1, actual_max_col + 1):
                value = self.get_cell_value(row, col)
                # 转换为字符串便于分析
                if value is None:
                    row_data.append("")
                elif isinstance(value, (int, float)):
                    row_data.append(f"[数值:{value}]")
                else:
                    row_data.append(str(value)[:50])  # 截断过长内容
            data.append(row_data)
        return data
    
    def get_merged_info(self) -> List[Dict]:
        """获取合并单元格信息"""
        merged_info = []
        for merged_range in self.ws.merged_cells.ranges:
            if merged_range.min_row <= 10:  # 只关注前10行
                merged_info.append({
                    'range': str(merged_range),
                    'rows': f"{merged_range.min_row}-{merged_range.max_row}",
                    'cols': f"{merged_range.min_col}-{merged_range.max_col}",
                    'value': str(self.ws.cell(merged_range.min_row, merged_range.min_col).value)[:30]
                })
        return merged_info
    
    def analyze_with_llm(self, 
                         llm_api_key: Optional[str] = None,
                         llm_base_url: Optional[str] = None,
                         llm_model: Optional[str] = None,
                         preview_max_rows: Optional[int] = None,
                         preview_max_cols: Optional[int] = None) -> HeaderAnalysis:
        """
        使用LLM直接分析Excel表头结构（包含行检测和列检测）
        
        参数:
            llm_api_key: LLM API密钥（必填）
            llm_base_url: LLM API地址（可选）
            llm_model: LLM模型名称（可选）
            preview_max_rows: 预览最大行数（可选，默认从配置读取）
            preview_max_cols: 预览最大列数（可选，默认从配置读取）
        
        返回:
            HeaderAnalysis 分析结果
        
        异常:
            如果LLM API Key未配置，抛出ValueError
        """
        # 检查API Key
        api_key = llm_api_key if llm_api_key is not None else EXCEL_LLM_API_KEY
        if not api_key:
            raise ValueError("LLM API Key 未配置，LLM分析是必需的")
        
        # 使用传入的参数或从配置读取默认值
        max_rows = preview_max_rows if preview_max_rows is not None else EXCEL_MAX_ROWS_PREVIEW
        max_cols = preview_max_cols if preview_max_cols is not None else EXCEL_MAX_COLS_PREVIEW
        
        preview_data = self.get_preview_data(max_rows=max_rows, max_cols=max_cols)
        merged_info = self.get_merged_info()
        
        # 获取列数信息
        max_col = self.ws.max_column
        
        # 构建分析提示词（包含行和列检测）
        prompt = self._build_analysis_prompt(preview_data, merged_info, max_col)
        
        # 调用LLM（使用传入的配置或从全局配置读取）
        result = self._call_llm(prompt, llm_api_key, llm_base_url, llm_model)
        
        if not result:
            raise ValueError("LLM调用失败，无法进行分析")
        
        # 解析LLM分析结果（包含行和列检测）
        analysis = self._parse_analysis_response(result)
        
        return analysis
    
    def _build_analysis_prompt(self, preview_data: List[List], merged_info: List[Dict], max_col: int) -> str:
        """构建LLM分析提示词（包含行检测和列检测）"""
        # 格式化预览数据为表格形式
        table_str = "行号 | 列1 | 列2 | 列3 | 列4 | 列5 | 列6 | 列7 | 列8 | ...\n" + "-" * 80 + "\n"
        for i, row in enumerate(preview_data, 1):
            row_str = " | ".join(str(cell)[:15] for cell in row[:8])
            table_str += f"  {i:2d}  | {row_str}\n"
        
        # 格式化合并单元格信息
        merged_str = "无" if not merged_info else "\n".join(
            f"  - {m['range']}: '{m['value']}'" for m in merged_info[:10]
        )
        
        prompt = f"""请分析以下Excel表格的结构，识别表头行、数据起始行和有效列。

【表格预览】（前{len(preview_data)}行，[数值:xxx]表示数值类型，空单元格显示为空）
{table_str}

【合并单元格信息】
{merged_str}

【表格信息】
- 总列数: {max_col}
- 总行数: {len(preview_data)}（预览）

请仔细分析表格结构，并以JSON格式返回分析结果：
{{
    "skip_rows": <需要跳过的无效行数（标题、注释等），从第1行开始计数>,
    "header_rows": <表头占用的行数>,
    "header_type": "<single或multi>",
    "data_start_row": <数据开始行（1-indexed）>,
    "valid_cols": [<有效列的索引列表，1-indexed，例如[1,2,3,5,7]表示第1,2,3,5,7列是有效的>],
    "confidence": "<high/medium/low>",
    "reason": "<分析说明：说明如何识别表头、数据起始行和有效列>"
}}

分析要点：
1. **行检测**：
   - 识别需要跳过的无效行（通常是标题、说明等，非空单元格很少的行）
   - 识别表头行（通常包含列名，可能是单行或多行）
   - 识别数据起始行（第一行包含实际数据的行，通常包含数值）

2. **列检测**：
   - 识别有效列：表头区域有内容或数据区域有数值数据的列
   - 过滤无效列：表头区域完全为空且数据区域完全为空或没有数值数据的列
   - valid_cols 应该是1-indexed的列索引列表，例如 [1,2,3,5,7] 表示第1,2,3,5,7列是有效的
   - 如果所有列都有效，valid_cols 可以为 null 或包含所有列索引

3. **表头类型**：
   - single: 单行表头
   - multi: 多行表头（合并单元格或分层结构）

4. **注意事项**：
   - skip_rows 是从第1行开始需要跳过的行数（例如skip_rows=2表示跳过第1-2行）
   - data_start_row 是数据开始的行号（1-indexed）
   - header_rows 是表头占用的行数
   - 确保 data_start_row = skip_rows + header_rows + 1
   - 只返回JSON，不要其他内容"""
        
        return prompt
    
    def _call_llm(self, prompt: str, llm_api_key: Optional[str] = None, 
                  llm_base_url: Optional[str] = None, llm_model: Optional[str] = None) -> str:
        """调用LLM API（支持OpenAI兼容接口）
        
        参数:
            prompt: 提示词
            llm_api_key: LLM API密钥（可选，如果不提供则从配置读取）
            llm_base_url: LLM API地址（可选，如果不提供则从配置读取）
            llm_model: LLM模型名称（可选，如果不提供则从配置读取）
        """
        # 优先使用传入的参数，否则从配置读取
        api_key = llm_api_key if llm_api_key is not None else EXCEL_LLM_API_KEY
        base_url = llm_base_url if llm_base_url is not None else EXCEL_LLM_BASE_URL
        model = llm_model if llm_model is not None else EXCEL_LLM_MODEL
        
        logger.info("=" * 60)
        logger.info("🤖 调用 LLM API 进行表头分析（包含行检测和列检测）")
        logger.info(f"🔗 EXCEL_LLM_BASE_URL: {base_url}")
        logger.info(f"📌 模型: {model}")
        logger.info(f"🔑 API Key: {'已配置' if api_key else '未配置'}")
        
        if not api_key:
            raise ValueError("LLM API Key 未配置，LLM分析是必需的")
            
        url = base_url
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        payload = {
            "model": model,
            "max_tokens": 1000,  # 增加token数量以支持列检测结果
            "messages": [{"role": "user", "content": prompt}]
        }
        
        logger.info(f"📡 发送 LLM API 请求到: {url}")
        logger.info(f"📝 提示词长度: {len(prompt)} 字符")
        
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=30)
            response.raise_for_status()
            result = response.json()
            llm_response = result['choices'][0]['message']['content']
            
            logger.info("✅ LLM API 调用成功")
            logger.info("=" * 60)
            logger.info("📝 LLM 响应内容:")
            logger.info("=" * 60)
            logger.info(llm_response)
            logger.info("=" * 60)
            
            return llm_response
        except Exception as e:
            logger.error(f"❌ LLM调用失败: {e}")
            logger.debug("异常详情:", exc_info=True)
            return None
    
    def _parse_analysis_response(self, response: str) -> HeaderAnalysis:
        """解析LLM分析结果（包含行检测和列检测）"""
        if not response:
            raise ValueError("LLM响应为空")
        
        try:
            # 提取JSON部分（支持嵌套JSON）
            # 先尝试找到第一个 { 到最后一个 } 之间的内容
            start_idx = response.find('{')
            end_idx = response.rfind('}')
            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                json_str = response[start_idx:end_idx + 1]
                data = json.loads(json_str)
            else:
                # 如果找不到完整的JSON，尝试用正则匹配
                json_match = re.search(r'\{.*\}', response, re.DOTALL)
                if not json_match:
                    raise ValueError("未找到JSON格式的响应")
                data = json.loads(json_match.group())
            
            # 解析行检测结果
            skip_rows = data.get('skip_rows', 0)
            header_rows = data.get('header_rows', 1)
            header_type = data.get('header_type', 'single')
            data_start_row = data.get('data_start_row', skip_rows + header_rows + 1)
            confidence = data.get('confidence', 'medium')
            reason = data.get('reason', 'LLM分析结果')
            
            # 解析列检测结果
            valid_cols = data.get('valid_cols', None)
            if valid_cols is None:
                # 如果为null，表示所有列都有效
                valid_cols = None
            elif isinstance(valid_cols, list):
                # 确保是整数列表
                valid_cols = [int(col) for col in valid_cols if isinstance(col, (int, str))]
                # 如果包含所有列，设为None
                max_col = self.ws.max_column
                if len(valid_cols) == max_col and set(valid_cols) == set(range(1, max_col + 1)):
                    valid_cols = None
            else:
                valid_cols = None
            
            # 验证数据起始行的一致性
            if data_start_row != skip_rows + header_rows + 1:
                logger.warning(f"⚠️ 数据起始行不一致，LLM返回: {data_start_row}，计算值: {skip_rows + header_rows + 1}，使用LLM返回的值")
            
            return HeaderAnalysis(
                skip_rows=skip_rows,
                header_rows=max(1, header_rows),
                header_type=header_type,
                data_start_row=data_start_row,
                confidence=confidence,
                reason=f"LLM分析: {reason}",
                valid_cols=valid_cols
            )
        except (json.JSONDecodeError, KeyError, ValueError) as e:
            logger.error(f"解析LLM分析响应失败: {e}")
            logger.error(f"响应内容: {response[:500]}")
            raise ValueError(f"解析LLM响应失败: {e}")
    
    
    def extract_headers(self, analysis: HeaderAnalysis) -> Tuple[List[str], Dict[str, Dict]]:
        """
        根据分析结果提取表头
        返回: (列名列表, 列结构元数据)
        """
        max_col = self.ws.max_column
        header_start = analysis.skip_rows + 1
        header_end = analysis.skip_rows + analysis.header_rows
        
        # 确定要处理的列（如果指定了有效列，只处理有效列）
        cols_to_process = analysis.valid_cols if analysis.valid_cols is not None else list(range(1, max_col + 1))
        
        logger.info(f"📋 提取表头: 处理 {len(cols_to_process)} 列")
        
        column_metadata = {}
        
        if analysis.header_type == 'single':
            # 单表头
            headers = []
            for col in cols_to_process:
                value = self.get_cell_value(header_start, col)
                col_name = str(value) if value else f'Column_{col}'
                headers.append(col_name)
                column_metadata[col_name] = {"level1": col_name}
            
            headers = self._handle_duplicate_names(headers)
            # 更新元数据的key
            column_metadata = {h: {"level1": h} for h in headers}
            return headers, column_metadata
        
        else:
            # 多表头：展平
            column_headers = []
            for col in cols_to_process:
                parts = []
                levels = {}
                for row_idx, row in enumerate(range(header_start, header_end + 1), 1):
                    value = self.get_cell_value(row, col)
                    if value is not None:
                        part = str(value).strip()
                        parts.append(part)
                        levels[f"level{row_idx}"] = part
                
                # 去重连续相同值
                unique_parts = []
                for p in parts:
                    if not unique_parts or p != unique_parts[-1]:
                        unique_parts.append(p)
                
                col_name = '_'.join(unique_parts) if unique_parts else f'Column_{col}'
                column_headers.append(col_name)
                column_metadata[col_name] = levels
            
            column_headers = self._handle_duplicate_names(column_headers)
            
            # 重新映射元数据
            new_metadata = {}
            for i, header in enumerate(column_headers):
                original_name = '_'.join(unique_parts) if (unique_parts := list(column_metadata.values())[i].values()) else f'Column_{i+1}'
                new_metadata[header] = list(column_metadata.values())[i]
            
            return column_headers, new_metadata
    
    def _handle_duplicate_names(self, names: List[str]) -> List[str]:
        """处理重复列名"""
        counts = defaultdict(int)
        result = []
        for name in names:
            if counts[name] > 0:
                result.append(f"{name}_{counts[name]}")
            else:
                result.append(name)
            counts[name] += 1
        return result
    
    def to_dataframe(self, analysis: HeaderAnalysis = None,
                    llm_api_key: Optional[str] = None,
                    llm_base_url: Optional[str] = None,
                    llm_model: Optional[str] = None,
                    preview_max_rows: Optional[int] = None,
                    preview_max_cols: Optional[int] = None) -> Tuple[pd.DataFrame, HeaderAnalysis, Dict[str, Dict]]:
        """
        转换为DataFrame
        
        参数:
            analysis: 预先的分析结果，如果为None则使用LLM自动分析（必选）
            llm_api_key: LLM API密钥（必填，如果analysis为None）
            llm_base_url: LLM API地址（可选）
            llm_model: LLM模型名称（可选）
            preview_max_rows: 预览最大行数（可选，默认从配置读取）
            preview_max_cols: 预览最大列数（可选，默认从配置读取）
        
        返回:
            (DataFrame, 分析结果, 列结构元数据)
        """
        if analysis is None:
            # 使用LLM进行分析（包含行检测和列检测）
            logger.info("🤖 使用LLM进行表头分析（包含行检测和列检测）...")
            analysis = self.analyze_with_llm(
                llm_api_key, 
                llm_base_url, 
                llm_model,
                preview_max_rows=preview_max_rows,
                preview_max_cols=preview_max_cols
            )
            logger.info("✅ LLM分析完成")
        
        headers, column_metadata = self.extract_headers(analysis)
        
        # 确定要读取的列（如果指定了有效列，只读取有效列）
        cols_to_read = analysis.valid_cols if analysis.valid_cols is not None else list(range(1, self.ws.max_column + 1))
        
        logger.info(f"📊 读取数据: 从 {len(cols_to_read)} 列读取数据")
        
        # 读取数据
        data = []
        for row in range(analysis.data_start_row, self.ws.max_row + 1):
            row_data = []
            for col in cols_to_read:
                row_data.append(self.ws.cell(row, col).value)
            if any(v is not None for v in row_data):
                data.append(row_data)
        
        df = pd.DataFrame(data, columns=headers)
        logger.info(f"✅ DataFrame 创建完成: {len(df)} 行 x {len(df.columns)} 列")
        return df, analysis, column_metadata
    
    def close(self):
        """关闭工作簿"""
        try:
            self.wb.close()
        except Exception:
            pass


def process_excel_file(
    filepath: str,
    output_dir: str,
    sheet_name: str = None,
    output_filename: str = None,
    llm_api_key: Optional[str] = None,
    llm_base_url: Optional[str] = None,
    llm_model: Optional[str] = None,
    preview_max_rows: Optional[int] = None,
    preview_max_cols: Optional[int] = None
) -> ExcelProcessResult:
    """
    处理Excel文件的主函数
    
    参数:
        filepath: Excel文件路径
        output_dir: 输出目录
        sheet_name: 工作表名称
        output_filename: 输出文件名（不含扩展名）
        llm_api_key: LLM API密钥（必填）
        llm_base_url: LLM API地址（可选）
        llm_model: LLM模型名称（可选）
        preview_max_rows: 预览最大行数（可选，默认从配置读取）
        preview_max_cols: 预览最大列数（可选，默认从配置读取）
    
    返回:
        ExcelProcessResult
    """
    try:
        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)
        
        # 处理Excel（使用LLM进行分析，包含行检测和列检测）
        processor = SmartHeaderProcessor(filepath, sheet_name)
        df, analysis, column_metadata = processor.to_dataframe(
            llm_api_key=llm_api_key,
            llm_base_url=llm_base_url,
            llm_model=llm_model,
            preview_max_rows=preview_max_rows,
            preview_max_cols=preview_max_cols
        )
        processor.close()
        
        # 生成输出文件名
        if not output_filename:
            base_name = Path(filepath).stem
            output_filename = f"{base_name}_processed"
        
        # 保存CSV
        csv_path = os.path.join(output_dir, f"{output_filename}.csv")
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        
        # 提取字段值样本（分组聚合后的常见值）
        logger.info("📊 提取字段值样本...")
        column_value_samples = extract_column_value_samples(df, max_samples_per_column=10)
        
        # 将值样本信息合并到列元数据中
        for col_name, samples in column_value_samples.items():
            if col_name in column_metadata:
                column_metadata[col_name]["value_samples"] = samples
            else:
                # 如果列不在元数据中（理论上不应该发生），创建新的元数据项
                column_metadata[col_name] = {"value_samples": samples}
        
        # 保存元数据
        metadata = {
            "header_analysis": analysis.to_dict(),
            "column_metadata": column_metadata,
            "column_names": list(df.columns),
            "row_count": len(df),
            "original_file": os.path.basename(filepath)
        }
        metadata_path = os.path.join(output_dir, f"{output_filename}_metadata.json")
        with open(metadata_path, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)
        
        # 打印处理后的JSON元数据
        logger.info("=" * 80)
        logger.info("📄 处理后的JSON元数据:")
        logger.info("=" * 80)
        logger.info(json.dumps(metadata, ensure_ascii=False, indent=2))
        logger.info("=" * 80)
        
        return ExcelProcessResult(
            success=True,
            header_analysis=analysis,
            processed_file_path=csv_path,
            metadata_file_path=metadata_path,
            column_names=list(df.columns),
            column_metadata=column_metadata,
            row_count=len(df),
            error_message=None
        )
        
    except Exception as e:
        import traceback
        error_msg = f"{str(e)}\n{traceback.format_exc()}"
        return ExcelProcessResult(
            success=False,
            header_analysis=None,
            processed_file_path=None,
            metadata_file_path=None,
            column_names=[],
            column_metadata={},
            row_count=0,
            error_message=error_msg
        )


def get_sheet_names(filepath: str) -> List[str]:
    """获取Excel文件的所有工作表名称"""
    try:
        wb = load_workbook(filepath, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception as e:
        return []


def extract_column_value_samples(
    df: pd.DataFrame,
    max_samples_per_column: int = 10,
    max_unique_ratio: float = 0.5
) -> Dict[str, Dict[str, Any]]:
    """
    提取每个字段的常见值样本（通过分组聚合）
    
    参数:
        df: 数据框
        max_samples_per_column: 每个字段最多保留的样本数量
        max_unique_ratio: 如果唯一值占比超过此比例，则只提供统计信息而不统计频率
    
    返回:
        字典，key为列名，value为包含常见值和统计信息的字典
    """
    column_samples = {}
    
    for col_name in df.columns:
        col_data = df[col_name]
        
        # 跳过完全为空的列
        if col_data.isna().all():
            continue
        
        # 计算非空值数量
        non_null_count = col_data.notna().sum()
        if non_null_count == 0:
            continue
        
        # 计算唯一值数量
        unique_count = col_data.nunique()
        unique_ratio = unique_count / non_null_count if non_null_count > 0 else 1.0
        
        sample_info = {
            "total_count": len(col_data),
            "non_null_count": int(non_null_count),
            "null_count": int(col_data.isna().sum()),
            "unique_count": int(unique_count),
            "data_type": str(col_data.dtype)
        }
        
        # 判断是否为数值类型
        is_numeric = pd.api.types.is_numeric_dtype(col_data)
        
        if is_numeric:
            # 数值类型：提供统计信息和常见值（如果唯一值不太多）
            sample_info["is_numeric"] = True
            non_null_data = col_data.dropna()
            if len(non_null_data) > 0:
                sample_info["min"] = float(non_null_data.min())
                sample_info["max"] = float(non_null_data.max())
                sample_info["mean"] = float(non_null_data.mean())
                sample_info["median"] = float(non_null_data.median())
            else:
                sample_info["min"] = None
                sample_info["max"] = None
                sample_info["mean"] = None
                sample_info["median"] = None
            
            # 如果唯一值不太多，也统计频率
            if unique_ratio <= max_unique_ratio and unique_count <= 100:
                value_counts = col_data.value_counts().head(max_samples_per_column)
                sample_info["top_values"] = [
                    {"value": float(k) if pd.notna(k) else None, "count": int(v)}
                    for k, v in value_counts.items()
                ]
            elif unique_count <= max_samples_per_column:
                # 即使唯一值比例高，但如果总数不多，也展示所有值
                value_counts = col_data.value_counts().head(max_samples_per_column)
                sample_info["top_values"] = [
                    {"value": float(k) if pd.notna(k) else None, "count": int(v)}
                    for k, v in value_counts.items()
                ]
                sample_info["note"] = f"唯一值较多（{unique_count}个），展示所有值"
        else:
            # 非数值类型：统计频率
            sample_info["is_numeric"] = False
            
            # 如果唯一值太多，只提供统计信息
            if unique_ratio > max_unique_ratio:
                sample_info["note"] = f"唯一值较多（{unique_count}个），仅展示部分常见值"
                # 仍然展示前N个最常见的值
                value_counts = col_data.value_counts().head(max_samples_per_column)
                sample_info["top_values"] = [
                    {"value": str(k) if pd.notna(k) else "空值", "count": int(v)}
                    for k, v in value_counts.items()
                ]
            else:
                # 唯一值不太多，统计所有值的频率
                value_counts = col_data.value_counts().head(max_samples_per_column)
                sample_info["top_values"] = [
                    {"value": str(k) if pd.notna(k) else "空值", "count": int(v)}
                    for k, v in value_counts.items()
                ]
        
        column_samples[col_name] = sample_info
    
    return column_samples


def _build_column_hierarchy_tree(column_metadata: Dict[str, Dict]) -> str:
    """
    构建列层级结构的树形展示
    
    参数:
        column_metadata: 列元数据字典
    
    返回:
        格式化的树形结构字符串
    """
    if not column_metadata:
        return ""
    
    # 构建树形结构
    tree = {}
    
    for col_name, meta in column_metadata.items():
        # 获取所有层级
        levels = []
        level_keys = sorted([k for k in meta.keys() if k.startswith('level')], 
                          key=lambda x: int(x.replace('level', '')))
        for level_key in level_keys:
            value = meta.get(level_key)
            if value and str(value).strip():
                levels.append(str(value).strip())
        
        # 如果没有层级信息，使用列名本身
        if not levels:
            levels = [col_name]
        
        # 构建树
        current = tree
        for i, level_value in enumerate(levels):
            if level_value not in current:
                current[level_value] = {}
            current = current[level_value]
    
    # 递归生成树形字符串
    def _format_tree(node: Dict, prefix: str = "", is_last: bool = True, depth: int = 0) -> List[str]:
        lines = []
        items = list(node.items())
        
        for idx, (key, children) in enumerate(items):
            is_last_item = (idx == len(items) - 1)
            current_prefix = "└─ " if is_last_item else "├─ "
            
            if children:
                # 有子节点
                lines.append(f"{prefix}{current_prefix}{key}")
                next_prefix = prefix + ("   " if is_last_item else "│  ")
                child_lines = _format_tree(children, next_prefix, is_last_item, depth + 1)
                lines.extend(child_lines)
            else:
                # 叶子节点
                lines.append(f"{prefix}{current_prefix}{key}")
        
        return lines
    
    tree_lines = _format_tree(tree)
    return "\n".join(tree_lines)


def generate_analysis_prompt(
    process_result: ExcelProcessResult,
    custom_prompt: str = None,
    include_metadata: bool = True
) -> str:
    """
    根据Excel处理结果生成数据分析提示词
    
    参数:
        process_result: Excel处理结果
        custom_prompt: 自定义分析提示词
        include_metadata: 是否包含列结构元数据
    
    返回:
        格式化的提示词
    """
    if not process_result.success:
        return ""
    
    # 基础信息
    prompt_parts = []
    
    # 添加语言要求（必须在最前面）
    prompt_parts.append("**重要要求：请使用中文进行所有分析和回答，包括代码注释、分析报告等所有内容。**")
    prompt_parts.append("")
    prompt_parts.append("**禁止要求：请不要生成任何图表绘制代码，包括但不限于：**")
    prompt_parts.append("- 不要使用 matplotlib、plotly、seaborn 等绘图库")
    prompt_parts.append("- 不要使用 plt.figure()、plt.plot()、plt.savefig() 等绘图函数")
    prompt_parts.append("- 不要使用 .plot()、.hist() 等 pandas 绘图方法")
    prompt_parts.append("- 不要保存任何图片文件（.png、.jpg、.svg 等）")
    prompt_parts.append("**请专注于数据分析和统计计算，不要生成可视化代码。**")
    prompt_parts.append("")
    
    if custom_prompt:
        prompt_parts.append(custom_prompt)
    else:
        prompt_parts.append("请对上传的数据进行全面分析，生成数据分析报告。")
    
    # 添加数据文件信息（重要：告诉AI需要读取CSV文件）
    if process_result.processed_file_path:
        csv_filename = os.path.basename(process_result.processed_file_path)
        prompt_parts.append(f"\n\n## 数据文件")
        prompt_parts.append(f"**重要：工作空间中已准备好处理后的CSV数据文件，文件名为：`{csv_filename}`**")
        prompt_parts.append(f"")
        prompt_parts.append(f"**请务必使用以下代码读取数据文件进行分析：**")
        prompt_parts.append(f"```python")
        prompt_parts.append(f"import pandas as pd")
        prompt_parts.append(f"")
        prompt_parts.append(f"# 读取处理后的CSV文件")
        prompt_parts.append(f"df = pd.read_csv('{csv_filename}')")
        prompt_parts.append(f"print(f'数据形状: {{df.shape}}')")
        prompt_parts.append(f"print(f'列名: {{list(df.columns)}}')")
        prompt_parts.append(f"```")
        prompt_parts.append(f"")
        prompt_parts.append(f"**注意：**")
        prompt_parts.append(f"- CSV文件已保存在当前工作空间目录中")
        prompt_parts.append(f"- 请使用 `pd.read_csv('{csv_filename}')` 读取数据")
        prompt_parts.append(f"- 不要仅根据元数据进行分析，必须读取实际数据文件进行计算")
        prompt_parts.append(f"")
    
    # 添加数据概况
    prompt_parts.append(f"\n## 数据概况")
    prompt_parts.append(f"- 数据行数: {process_result.row_count}")
    prompt_parts.append(f"- 列数: {len(process_result.column_names)}")
    
    # 添加表头类型信息（仅保留对分析有用的信息）
    if process_result.header_analysis:
        ha = process_result.header_analysis
        if ha.header_type == 'multi':
            prompt_parts.append(f"\n## 表头结构")
            prompt_parts.append(f"- 表头类型: 多级表头（{ha.header_rows}层）")
    
    # 添加列结构元数据（帮助AI理解列之间的关系）
    if include_metadata and process_result.column_metadata:
        # 检查是否有多级结构
        has_multi_level = any(
            len(meta) > 1 
            for meta in process_result.column_metadata.values()
        )
        
        if has_multi_level:
            prompt_parts.append(f"\n## 列层级结构（多级表头语义关系）")
            prompt_parts.append("以下树形结构展示了列之间的层级分组关系，有助于理解数据的业务含义：")
            prompt_parts.append("")
            hierarchy_tree = _build_column_hierarchy_tree(process_result.column_metadata)
            if hierarchy_tree:
                prompt_parts.append(hierarchy_tree)
            else:
                # 如果树形构建失败，使用分组展示
                groups = defaultdict(list)
                for col_name, meta in process_result.column_metadata.items():
                    level1 = meta.get('level1', col_name)
                    groups[level1].append(col_name)
                
                for group, cols in groups.items():
                    if len(cols) > 1:
                        prompt_parts.append(f"- {group}: {', '.join(cols)}")
    
    # 添加完整的列名列表
    prompt_parts.append(f"\n## 完整列名列表")
    if len(process_result.column_names) <= 30:
        # 如果列数不多，全部展示
        for idx, col_name in enumerate(process_result.column_names, 1):
            prompt_parts.append(f"{idx}. {col_name}")
    else:
        # 如果列数很多，展示前20个和后10个
        for idx, col_name in enumerate(process_result.column_names[:20], 1):
            prompt_parts.append(f"{idx}. {col_name}")
        prompt_parts.append(f"... (省略中间 {len(process_result.column_names) - 30} 列) ...")
        for idx, col_name in enumerate(process_result.column_names[-10:], len(process_result.column_names) - 9):
            prompt_parts.append(f"{idx}. {col_name}")
        prompt_parts.append(f"\n(共 {len(process_result.column_names)} 列)")
    
    # 添加字段值样本信息（以JSON格式提供，更结构化）
    if include_metadata and process_result.column_metadata:
        prompt_parts.append(f"\n## 字段值样本（常见值统计）")
        prompt_parts.append("以下JSON格式展示了每个字段的常见值及其出现频率，有助于理解数据的实际内容：")
        prompt_parts.append("")
        
        # 构建包含值样本的column_metadata JSON
        column_metadata_with_samples = {}
        for col_name in process_result.column_names:
            if col_name in process_result.column_metadata:
                column_metadata_with_samples[col_name] = process_result.column_metadata[col_name]
        
        # 将column_metadata转换为格式化的JSON字符串
        prompt_parts.append("```json")
        prompt_parts.append(json.dumps(column_metadata_with_samples, ensure_ascii=False, indent=2))
        prompt_parts.append("```")
        prompt_parts.append("")
        
        prompt_parts.append("**说明：**")
        prompt_parts.append("- 每个字段的元数据包含 `value_samples` 字段，其中包含该字段的统计信息和常见值")
        prompt_parts.append("- `value_samples.top_values` 数组展示了出现频率最高的值及其出现次数")
        prompt_parts.append("- 对于数值类型字段，还包含 `min`、`max`、`mean`、`median` 等统计信息")
    
    # 在末尾再次强调要求
    prompt_parts.append("\n\n**再次提醒：请务必使用中文进行所有分析、代码注释和报告撰写，且不要生成任何图表绘制代码。**")
    
    full_prompt = '\n'.join(prompt_parts)
    
    # 打印生成的提示词
    logger.info("=" * 80)
    logger.info("📝 生成的AI分析提示词:")
    logger.info("=" * 80)
    logger.info(full_prompt)
    logger.info("=" * 80)
    
    return full_prompt

